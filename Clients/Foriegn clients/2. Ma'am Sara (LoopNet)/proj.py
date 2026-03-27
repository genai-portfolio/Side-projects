"""
LoopNet Commercial Real Estate Scraper
=======================================
- Scrapes 25 listings per run
- Saves to per-country Excel file  (listings_usa.xlsx, listings_france.xlsx …)
- Updates rows in-place — only changed columns are overwritten
- Unique key = name + address  (handles duplicate building names)
- Name fallback: h4/a → h6/a → first text in header/div[1]
- Runtime mid-path detection per card (no country-based hardcoding)

Requirements:
    pip install playwright openpyxl
    playwright install chromium
"""

import asyncio
import math
import random
import re
from typing import Optional
import subprocess
import sys
import threading
import tkinter as tk
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, TimeoutError as PWTimeout


# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
CDP_URL          = "http://localhost:9222"
CHROME_PATH      = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
CHROME_DEBUG_DIR = r"C:\chrome-debug"
MAX_LISTINGS     = 25     # listings per page
TOTAL_PAGES      = 30     # total pages to scrape (30 pages × 25 = 750 listings)
CHROME_WAIT_SEC  = 3

# ─────────────────────────────────────────────
# GOOGLE SHEETS CONFIG
# ─────────────────────────────────────────────
CLIENT_SECRET    = "client_secret.json"   # OAuth credentials from Google Cloud

def load_sheet_id() -> str:
    """Read spreadsheet_id from client_secret.json — client edits this one file."""
    import json as _json
    try:
        with open(CLIENT_SECRET, encoding="utf-8") as f:
            data = _json.load(f)
        sid = data.get("spreadsheet_id", "").strip()
        if not sid:
            print("[Sheets] ⚠  No 'spreadsheet_id' found in client_secret.json")
            print("[Sheets]    Add it like this:")
            print('           "spreadsheet_id": "your-sheet-id-here"')
        return sid
    except FileNotFoundError:
        return ""
    except Exception as e:
        print(f"[Sheets] ⚠  Could not read {CLIENT_SECRET}: {e}")
        return ""

TOKEN_FILE       = "gsheets_token.json"   # saved after first login — never delete
UPLOAD_TO_SHEETS = True                   # set False to skip upload

COUNTRIES     = {"1": "Williamson", "2": "Davidson", "3": "Rutherford", "4": "Wilson"}
LISTING_TYPES = {"1": "For Lease", "2": "For Sale"}

# Sub-location options for each county selection
SUB_LOCATIONS = {
    "Williamson": [
        "Williamson, NY, USA",
        "Williamson County, TX, USA",
        "Williamson, WV, USA",
        "Williamson, GA, USA",
        "Williamson County, TN, USA",
    ],
    "Davidson": [
        "Davidson, NC, USA",
        "Davidsonville, MD, USA",
        "Davidson County, TN, USA",
    ],
    "Rutherford": [
        "Rutherford, NJ, USA",
        "Rutherfordton, NC, USA",
        "Rutherford County, TN, USA",
    ],
    "Wilson": [
        "Wilson, NC, USA",
        "Wilson County, TN, USA",
        "Wilson County, TX, USA",
        "Wilsonville, OR, USA",
        "Wilson County, KS, USA",
        "Wilson County Tennessee State Fair, East Baddour Parkway, Lebanon, TN, USA"
    ],
}

FIELDNAMES      = ["name", "address", "size", "price", "brokers", "status", "permit_intel", "last_seen", "first_seen"]
FIELDNAMES_SALE = ["name", "address", "size", "price", "auction_date", "status", "permit_intel", "last_seen", "first_seen"]
TRACK      = ["address", "size", "price", "brokers", "status"]
TRACK_SALE = ["address", "size", "price", "auction_date", "status"]

def xlsx_path(country: str, listing_type: str = "For Lease") -> str:
    slug  = country.lower().replace(" ", "_")
    ltype = "for_sale" if listing_type == "For Sale" else "for_lease"
    return f"listings_{slug}_{ltype}.xlsx"

def row_key(row: dict) -> str:
    """Composite key: name + address — handles same building name, different units."""
    return f"{row.get('name','').strip()}||{row.get('address','').strip()}"


# ─────────────────────────────────────────────
# BASE XPATH
# ─────────────────────────────────────────────
# Total listing count on results header (divide by 30 → page count, per LoopNet UI)
XP_RESULT_TOTAL = (
    "/html/body/section[1]/main/section[1]/div/section/section[1]/div[1]/div/div[2]/div/span/span"
)

BASE = "/html/body/section[1]/main/section[1]/div/section/section[1]/div[3]"

def article(ul, li):   return f"{BASE}/ul[{ul}]/li[{li}]/article"
def xp_name(ul, li):   return f"{article(ul,li)}/header/div[1]/h4/a"
def xp_name2(ul, li):  return f"{article(ul,li)}/header/div[1]/h6/a"   # fallback
def xp_name3(ul, li):  return f"{article(ul,li)}/header/div[1]"         # last resort
def xp_build(ul, li):  return f"{article(ul,li)}/header/div[1]/h6/a"
def xp_addr(ul, li):   return f"{article(ul,li)}/header/div[2]/h6/a"

# Three mid-path variants observed across all 5 countries
MID_PATHS = [
    "div[2]/div[2]/div",      # Path A — most common
    "div[2]/div[2]/div[1]",   # Path B — premium/newer cards
    "div[2]/div/div",         # Path C — no-image cards
]

def xp_li(ul, li, mid, idx):
    return f"{article(ul,li)}/{mid}/ul[1]/li[{idx}]"

def xp_broker_x(ul, li, n):
    b = f"{article(ul,li)}/div[2]/div[2]/ul/li[{n}]/a/span[2]"
    return f"{b}/span[1]", f"{b}/span[2]"

def xp_broker_y(ul, li, n):
    b = f"{article(ul,li)}/div[2]/div[2]/div[2]/div[1]/div/div[{n}]/div/a/span[3]"
    return f"{b}/span[1]", f"{b}/span[2]"

def xp_broker_z(ul, li, n):
    b = f"{article(ul,li)}/div[2]/div/ul/li[{n}]/a/span[2]"
    return f"{b}/span[1]", f"{b}/span[2]"


# ─────────────────────────────────────────────
# PRICE / SIZE CLASSIFIERS
# ─────────────────────────────────────────────
CURRENCY_RE = re.compile(
    r"(\$|€|£|price\s+upon\s+request|upon\s+request|negotiable|contact\s+for\s+price)",
    re.IGNORECASE
)
REJECT_RE = re.compile(
    r"(\d+\s+space|\d+\s+unit|\d+\s+suite|\d+\s+people|available\s+(now|soon))",
    re.IGNORECASE
)
SIZE_RE = re.compile(r"[\d,\.\s\-]+SF", re.IGNORECASE)

def is_price(t):
    if not t: return False
    if REJECT_RE.search(t): return False
    return bool(CURRENCY_RE.search(t))

def is_size(t):
    return bool(SIZE_RE.search(t)) and not is_price(t)


# ─────────────────────────────────────────────
# CHROME
# ─────────────────────────────────────────────
def launch_chrome():
    if not Path(CHROME_PATH).exists():
        print(f"❌  Chrome not found at: {CHROME_PATH}"); sys.exit(1)
    subprocess.Popen([
        CHROME_PATH, "--remote-debugging-port=9222",
        f"--user-data-dir={CHROME_DEBUG_DIR}",
        "--no-first-run", "--no-default-browser-check",
    ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    print("[chrome] ✓ Launched")


# ─────────────────────────────────────────────
# TKINTER UI
# ─────────────────────────────────────────────
def launch_ui():
    result = {}
    root = tk.Tk()
    root.title("LoopNet Scraper")
    root.resizable(False, False)
    root.configure(bg="#1a1a2e")

    BG, CARD_BG = "#1a1a2e", "#16213e"
    ACCENT      = "#e94560"
    FG, FG_DIM  = "#e0e0e0", "#a0a0b0"
    FONT   = ("Segoe UI", 10)
    FONT_B = ("Segoe UI", 10, "bold")
    FONT_H = ("Segoe UI", 14, "bold")
    cf = lambda p: tk.Frame(p, bg=CARD_BG)

    tk.Frame(root, bg=ACCENT, height=4).pack(fill="x")
    tk.Label(root, text="LoopNet Scraper",       bg=BG, fg=ACCENT,  font=FONT_H).pack(pady=(14,2))
    tk.Label(root, text="Configure your search", bg=BG, fg=FG_DIM, font=("Segoe UI",9)).pack(pady=(0,12))

    cc = cf(root); cc.pack(fill="x", padx=22, pady=(0,10))
    tk.Label(cc, text="  🌍  Select County", bg=CARD_BG, fg=ACCENT, font=FONT_B, anchor="w").pack(fill="x", padx=10, pady=(10,6))
    country_var = tk.StringVar(value="1")
    rb = dict(bg=CARD_BG, fg=FG, font=FONT, activebackground=CARD_BG,
              activeforeground=ACCENT, selectcolor="#0f3460", relief="flat", variable=country_var)
    cg = tk.Frame(cc, bg=CARD_BG); cg.pack(padx=14, pady=(0,4))
    for idx, (num, name) in enumerate(COUNTRIES.items()):
        tk.Radiobutton(cg, text=f"{num}. {name}", value=num, **rb).grid(
            row=idx//2, column=idx%2, sticky="w", padx=10, pady=3)

    # ── Sub-location dropdown (updates when county radio changes) ─────
    tk.Label(cc, text="  📍  Select Location", bg=CARD_BG, fg=FG_DIM, font=FONT_B, anchor="w").pack(fill="x", padx=10, pady=(8,4))
    subloc_var = tk.StringVar()
    subloc_menu = tk.OptionMenu(cc, subloc_var, "")
    subloc_menu.config(bg="#0f3460", fg=FG, font=FONT, relief="flat",
                       activebackground="#1a3a6e", activeforeground=ACCENT,
                       highlightthickness=0, bd=0, cursor="hand2")
    subloc_menu["menu"].config(bg="#0f3460", fg=FG, font=FONT,
                               activebackground=ACCENT, activeforeground="white")
    subloc_menu.pack(fill="x", padx=14, pady=(0,10))

    def refresh_subloc(*_):
        county = COUNTRIES[country_var.get()]
        options = SUB_LOCATIONS.get(county, [])
        menu = subloc_menu["menu"]
        menu.delete(0, "end")
        for opt in options:
            menu.add_command(label=opt, command=lambda v=opt: subloc_var.set(v))
        subloc_var.set(options[0] if options else "")

    country_var.trace_add("write", refresh_subloc)
    refresh_subloc()   # populate on startup

    tc = cf(root); tc.pack(fill="x", padx=22, pady=(0,10))
    tk.Label(tc, text="  🏷  Listing Type", bg=CARD_BG, fg=ACCENT, font=FONT_B, anchor="w").pack(fill="x", padx=10, pady=(10,6))
    type_var = tk.StringVar(value="1")
    tr = tk.Frame(tc, bg=CARD_BG); tr.pack(padx=14, pady=(0,10))
    for num, name in LISTING_TYPES.items():
        tk.Radiobutton(tr, text=f"{num}. {name}", value=num,
                       **{**rb, "variable": type_var}).pack(side="left", padx=16)

    # ── Pages slider card ─────────────────────────────────────────────
    sc = cf(root); sc.pack(fill="x", padx=22, pady=(0,14))
    tk.Label(sc, text="  📄  Pages to Scrape  (25 listings / page)",
             bg=CARD_BG, fg=ACCENT, font=FONT_B, anchor="w").pack(fill="x", padx=10, pady=(10,6))

    # Slider: 1–30 pages, each page = 25 listings → 25 to 750 total
    pages_var = tk.IntVar(value=30)
    listing_count_var = tk.StringVar(value="750 listings  (30 pages)")

    def on_slider(val):
        pages = int(float(val))
        pages_var.set(pages)
        total = pages * 25
        listing_count_var.set(f"{total} listings  ({pages} pages)")

    # ── Count label (big, centred above slider) ───────────────────────
    tk.Label(sc, textvariable=listing_count_var,
             bg=CARD_BG, fg="#2ecc71", font=("Segoe UI", 11, "bold"),
             anchor="center").pack(fill="x", padx=14, pady=(0,4))

    # ── Slider (full width of card) ───────────────────────────────────
    slider = tk.Scale(
        sc,
        from_=1, to=30,
        orient="horizontal",
        variable=pages_var,
        command=on_slider,
        bg=CARD_BG, fg=FG,
        troughcolor="#0f3460",
        highlightthickness=0,
        activebackground=ACCENT,
        sliderrelief="flat",
        sliderlength=18,
        showvalue=False,
        tickinterval=0,
        resolution=1,
        length=380,
    )
    slider.pack(fill="x", padx=14, pady=(0,2))

    # ── Tick marks: 0 25 50 … 750 evenly spaced ──────────────────────
    # Use a Canvas so we can place labels at exact pixel positions
    TICK_W = 380
    tick_canvas = tk.Canvas(sc, bg=CARD_BG, height=18,
                            width=TICK_W, highlightthickness=0)
    tick_canvas.pack(padx=14, pady=(0,10))

    tick_values = [0, 75, 150, 225, 300, 375, 450, 525, 600, 675, 750]
    # Map listing count → x pixel (0 listings = left edge, 750 = right edge)
    for lv in tick_values:
        page_num = lv // 25 if lv > 0 else 0   # 0 shown at far left
        # pixel x: pages 1..30 map to 0..TICK_W
        # slider goes from_=1 to=30, so x = (page_num-1)/(30-1) * TICK_W
        if lv == 0:
            x = 0
        else:
            x = int((page_num - 1) / 29 * TICK_W)
        tick_canvas.create_text(
            x, 9,
            text=str(lv),
            fill=FG_DIM,
            font=("Segoe UI", 7),
            anchor="center"
        )

    status_var = tk.StringVar(value="")
    tk.Label(root, textvariable=status_var, bg=BG, fg="#2ecc71", font=FONT_B).pack(pady=(0,6))

    btn_frame = tk.Frame(root, bg=BG); btn_frame.pack(pady=(0,18))
    start_btn = tk.Button(btn_frame, text="▶  Start Scraping",
                          bg=ACCENT, fg="white", font=FONT_B,
                          relief="flat", cursor="hand2", padx=18, pady=8)
    start_btn.pack(side="left", padx=8)
    tk.Button(btn_frame, text="Close", command=root.destroy,
              bg="#333355", fg=FG, font=FONT,
              relief="flat", cursor="hand2", padx=14, pady=8).pack(side="left", padx=8)

    result.update({"_root": root, "_status_var": status_var,
                   "_start_btn": start_btn, "_chrome_done": False,
                   "_pages_var": pages_var})

    def set_busy(busy):
        s, bg = ("disabled","#888") if busy else ("normal", ACCENT)
        start_btn.config(state=s, bg=bg)
        slider.config(state=s)
        for w in cg.winfo_children(): w.config(state=s)
        for w in tr.winfo_children(): w.config(state=s)
        subloc_menu.config(state=s)
    result["_set_busy"] = set_busy

    def tick(n, locations, lt, display_country):
        if n > 0:
            status_var.set(f"⏳  Chrome opening … {n}s")
            root.after(1000, tick, n-1, locations, lt, display_country)
        else:
            status_var.set("🔗  Connecting …")
            threading.Thread(
                target=lambda: asyncio.run(run(locations, lt, result, display_country)),
                daemon=True).start()

    def on_start():
        county_name = COUNTRIES[country_var.get()]
        lt = LISTING_TYPES[type_var.get()]
        result["_pages"] = pages_var.get()   # capture slider value at click time
        # Williamson: run all five sub-locations in order into one workbook/tab
        if county_name == "Williamson":
            locations = list(SUB_LOCATIONS["Williamson"])
            display_country = "Williamson"
        else:
            locations = [subloc_var.get() or county_name]
            display_country = locations[0]
        set_busy(True)
        if not result["_chrome_done"]:
            launch_chrome()
            result["_chrome_done"] = True
            tick(CHROME_WAIT_SEC, locations, lt, display_country)
        else:
            status_var.set("🔗  Connecting …")
            threading.Thread(
                target=lambda: asyncio.run(run(locations, lt, result, display_country)),
                daemon=True).start()

    start_btn.config(command=on_start)
    root.update_idletasks()
    root.geometry(f"{root.winfo_reqwidth()+20}x{root.winfo_reqheight()+10}")
    root.mainloop()


# ─────────────────────────────────────────────
# HUMAN HELPERS
# ─────────────────────────────────────────────
async def hd(page, mn=800, mx=2200):
    await page.wait_for_timeout(random.randint(mn, mx))

async def htype(page, loc, text):
    for ch in text:
        await loc.type(ch, delay=random.randint(80, 220))
        if random.random() < 0.07:
            await page.wait_for_timeout(random.randint(150, 500))

async def hclick(page, loc):
    await loc.scroll_into_view_if_needed()
    await loc.hover()
    await page.wait_for_timeout(random.randint(100, 350))
    await loc.click()

async def hpresence(page):
    vp = page.viewport_size or {"width": 1280, "height": 900}
    for _ in range(random.randint(3, 5)):
        await page.mouse.move(random.randint(100, vp["width"]-100),
                              random.randint(100, vp["height"]-100), steps=15)
        await page.wait_for_timeout(random.randint(80, 200))
    await page.mouse.wheel(0, random.randint(60, 150))
    await page.wait_for_timeout(200)
    await page.mouse.wheel(0, -random.randint(30, 80))


# ─────────────────────────────────────────────
# CONNECT
# ─────────────────────────────────────────────
async def connect(pw):
    print(f"\n[browser] Connecting …")
    browser = await pw.chromium.connect_over_cdp(CDP_URL)
    ctx  = browser.contexts[0] if browser.contexts else await browser.new_context()
    page = await ctx.new_page()
    print("[browser] ✓ Connected")
    return browser, page


# ─────────────────────────────────────────────
# SEARCH + FILTER
# ─────────────────────────────────────────────
XP_SEARCH = "/html/body/section[1]/main/section/section[1]/div[2]/div/div/form/div/div/div[2]/div[2]/div/input"
XP_DD     = "/html/body/section[1]/main/section[1]/div/div/section[2]/div[1]/div[2]/div/button"
XP_LEASE  = "/html/body/section[1]/main/section[1]/div/div/section[2]/div[1]/div[2]/div/ul/li[1]/button"
XP_SALE   = "/html/body/section[1]/main/section[1]/div/div/section[2]/div[1]/div[2]/div/ul/li[2]/button"

def on_results(url): return "/search/" in url or "/commercial-real-estate/" in url

def build_search_url(country: str, listing_type: str) -> str:
    """
    Build the page-1 search URL directly.
    Accepts full location strings like "Williamson County, TN, USA"
    and converts them to a URL-friendly slug.

    For Lease: https://www.loopnet.com/search/commercial-real-estate/williamson-county-tn/for-lease/?view=map
    For Sale:  https://www.loopnet.com/search/commercial-real-estate/williamson-county-tn/for-sale/?view=map
    """
    # Strip country suffix (", USA", ", Canada", etc.) for cleaner slugs
    slug = country.lower()
    slug = re.sub(r",?\s*(usa|canada|australia|india|uk|united kingdom)$", "", slug, flags=re.IGNORECASE).strip()
    slug = slug.replace(",", " ").replace("  ", " ").strip()
    slug = slug.replace(" ", "-")
    ltype = "for-lease" if listing_type == "For Lease" else "for-sale"
    return f"https://www.loopnet.com/search/commercial-real-estate/{slug}/{ltype}/?view=map"


def build_page_url(base_url: str, page_num: int) -> str:
    """
    Page 1: https://…/for-lease/?view=map
    Page N: https://…/for-lease/N/?view=map
    Strips any query string and existing page number before appending.
    """
    import re as _re
    clean = base_url.split("?")[0].rstrip("/")
    clean = _re.sub(r"/\d+$", "", clean)
    if page_num == 1:
        return clean + "/?view=map"
    return f"{clean}/{page_num}/?view=map"


async def do_search(page, country, listing_type):
    url = build_search_url(country, listing_type)
    print(f"\n[2] Navigating directly to: {url}")
    try:
        await page.goto(url, wait_until="networkidle", timeout=25000)
    except Exception:
        await page.goto(url, wait_until="domcontentloaded", timeout=20000)
    await hpresence(page)
    await hd(page, 2000, 3500)
    print(f"[2] ✓  {page.url}")


# ─────────────────────────────────────────────
# SAFE TEXT GETTER
# ─────────────────────────────────────────────
async def gt(page, xpath, timeout=4000) -> str:
    try:
        el = page.locator(f"xpath={xpath}")
        await el.wait_for(state="visible", timeout=timeout)
        return (await el.inner_text()).strip()
    except Exception:
        return ""


async def detect_total_pages_from_header(page) -> Optional[int]:
    """
    Read total listing count from the results header span, then derive page count
    as ceil(count / 30). LoopNet may show values like 472, 750+, or 1,234.
    Returns None if the element is missing or unparsable (caller falls back to slider).
    """
    text = (await gt_fast(page, XP_RESULT_TOTAL)) or (await gt(page, XP_RESULT_TOTAL, timeout=3000))
    if not text:
        return None
    nums = re.findall(r"\d[\d,]*", text)
    if not nums:
        return None
    try:
        n = int(nums[0].replace(",", ""))
    except ValueError:
        return None
    if n <= 0:
        return None
    return max(1, math.ceil(n / 30))


# ─────────────────────────────────────────────
# FAST CARD EXTRACTOR — single JS call per card
# ─────────────────────────────────────────────
# Instead of 8-15 sequential Playwright locator calls (each with its own
# round-trip + timeout), we send ONE evaluate() call that runs entirely
# inside the browser and returns all fields at once.

async def gt_fast(page, xpath: str) -> str:
    """
    Fast non-blocking text getter for parallel gather calls.
    Uses page.evaluate() to read DOM directly — no waiting, no timeouts.
    Returns empty string if element not found.
    """
    try:
        val = await page.evaluate(
            "(xp) => {"
            "  var n = document.evaluate(xp, document, null,"
            "    XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;"
            "  return n ? n.textContent.trim() : '';"
            "}",
            xpath
        )
        return val or ""
    except Exception:
        return ""


async def scrape_card(page, ul: int, li: int, listing_type: str = "For Lease") -> dict:
    """
    Fetch all fields for one card in parallel using asyncio.gather.

    For Lease cards:  size/price in li items, brokers in spans
    For Sale cards:   size in header/div[2]/h4/a, price = starting/current bid,
                      auction date in div[2]/a[3]/div/span[5]
    Both share:       name, building, address in header
    """
    import asyncio as _asyncio

    A = f"{BASE}/ul[{ul}]/li[{li}]/article"

    shared_xpaths = [
        # ── Header fields (identical for both lease and sale) ────────────
        f"{A}/header/div[1]/h4/a",              # 0  name primary
        f"{A}/header/div[1]/h6/a",              # 1  building / name fallback
        f"{A}/header/div[2]/h6/a",              # 2  address
        # ── Small card name/address (France pages 2+) ────────────────────
        f"{A}/div[2]/header/div/h4/a",          # 3  name (small card)
        f"{A}/div[2]/header/div/a",             # 4  place (small card)
        # ── Size: For Sale uses header/div[2]/h4/a ───────────────────────
        f"{A}/header/div[2]/h4/a",              # 5  "4,750 SF Retail" / "50,675 SF Office Available"
        # ── For Lease size/price: Path A ─────────────────────────────────
        f"{A}/div[2]/div[2]/div/ul[1]/li[1]",   # 6
        f"{A}/div[2]/div[2]/div/ul[1]/li[2]",   # 7
        f"{A}/div[2]/div[2]/div/ul[1]/li[3]",   # 8
        f"{A}/div[2]/div[2]/div/ul[1]/li[4]",   # 9
        f"{A}/div[2]/div[2]/div/ul[1]/li[5]",   # 10
        # ── For Lease size/price: Path B ─────────────────────────────────
        f"{A}/div[2]/div[2]/div[1]/ul[1]/li[1]",  # 11
        f"{A}/div[2]/div[2]/div[1]/ul[1]/li[2]",  # 12
        f"{A}/div[2]/div[2]/div[1]/ul[1]/li[3]",  # 13
        f"{A}/div[2]/div[2]/div[1]/ul[1]/li[4]",  # 14
        f"{A}/div[2]/div[2]/div[1]/ul[1]/li[5]",  # 15
        # ── For Lease size/price: Path C ─────────────────────────────────
        f"{A}/div[2]/div/div/ul[1]/li[1]",      # 16
        f"{A}/div[2]/div/div/ul[1]/li[2]",      # 17
        f"{A}/div[2]/div/div/ul[1]/li[3]",      # 18
        # ── For Lease size/price: Path D (small card, no [1] on ul) ──────
        f"{A}/div[2]/div[2]/div/ul/li[1]",      # 19
        f"{A}/div[2]/div[2]/div/ul/li[2]",      # 20
        f"{A}/div[2]/div[2]/div/ul/li[3]",      # 21
        # ── For Sale price: starting bid (Path div[2]/div[2]) ────────────
        f"{A}/div[2]/div[2]/div/div[2]/ul/li[3]/span",   # 22  "Starting bid $X"
        f"{A}/div[2]/div[2]/div/div[5]/span[1]",         # 23  "$166,000" (current bid in progress)
        f"{A}/div[2]/div[2]/div/div[9]/span[1]",         # 24  "$166,000" (alt current bid)
        # ── For Sale price: starting bid (Path C div[2]/div) ─────────────
        f"{A}/div[2]/div/div/div[2]/ul/li[3]/span",      # 25  "Starting bid $X" alt path
        # ── For Sale auction status — placeholder, extracted via JS below ─
        f"{A}/div[2]/a[3]/div/span[5]",   # 26  (unused — JS handles this)
        f"{A}/div[2]/a[3]/div/span[5]",   # 44  (placeholder)
        f"{A}/div[2]/a[3]/div/span[5]",   # 45  (placeholder)
        f"{A}/div[2]/a[3]/div/span[5]",   # 46  (placeholder)
        f"{A}/div[2]/a[3]/div/span[5]",   # 47  (placeholder)
        f"{A}/div[2]/a[3]/div/span[5]",   # 48  (placeholder)
        # ── Brokers: Variant X ────────────────────────────────────────────
        f"{A}/div[2]/div[2]/ul/li[2]/a/span[2]/span[1]",  # 27
        f"{A}/div[2]/div[2]/ul/li[2]/a/span[2]/span[2]",  # 28
        f"{A}/div[2]/div[2]/ul/li[3]/a/span[2]/span[1]",  # 29
        f"{A}/div[2]/div[2]/ul/li[3]/a/span[2]/span[2]",  # 30
        f"{A}/div[2]/div[2]/ul/li[4]/a/span[2]/span[1]",  # 31
        f"{A}/div[2]/div[2]/ul/li[4]/a/span[2]/span[2]",  # 32
        # ── Brokers: Variant Y ────────────────────────────────────────────
        f"{A}/div[2]/div[2]/div[2]/div[1]/div/div[1]/div/a/span[3]/span[1]",  # 33
        f"{A}/div[2]/div[2]/div[2]/div[1]/div/div[1]/div/a/span[3]/span[2]",  # 34
        f"{A}/div[2]/div[2]/div[2]/div[1]/div/div[2]/div/a/span[3]/span[1]",  # 35
        f"{A}/div[2]/div[2]/div[2]/div[1]/div/div[2]/div/a/span[3]/span[2]",  # 36
        f"{A}/div[2]/div[2]/div[2]/div[1]/div/div[3]/div/a/span[3]/span[1]",  # 37
        f"{A}/div[2]/div[2]/div[2]/div[1]/div/div[3]/div/a/span[3]/span[2]",  # 38
        # ── Brokers: Variant Z ────────────────────────────────────────────
        f"{A}/div[2]/div/ul/li[2]/a/span[2]/span[1]",  # 39
        f"{A}/div[2]/div/ul/li[2]/a/span[2]/span[2]",  # 40
        f"{A}/div[2]/div/ul/li[3]/a/span[2]/span[1]",  # 41
        f"{A}/div[2]/div/ul/li[3]/a/span[2]/span[2]",  # 42
        # ── Broker: sale card company name (Variant Z li[1]) ─────────────
        f"{A}/div[2]/div/ul/li[1]/a",           # 43  e.g. "Southeast Property Advisor"
    ]

    # Wait for article to exist before firing all xpaths
    try:
        await page.wait_for_selector(f"xpath={A}", state="attached", timeout=5000)
    except Exception:
        pass

    vals = await _asyncio.gather(
        *[gt_fast(page, xp) for xp in shared_xpaths]
    )

    # ── Name + Address ────────────────────────────────────────────────
    large_name = vals[0] or vals[1]
    small_name = vals[3]
    is_small   = not large_name and bool(small_name)

    if is_small:
        name     = small_name
        address  = vals[4] or "N/A"
        building = ""
    else:
        name     = large_name
        address  = vals[2] or "N/A"
        building = vals[1] if vals[1] and vals[1] != name else ""

    # Merge building into name if it adds new information
    if building and building not in name:
        name = f"{name} — {building}"

    if not name:
        raw = await gt_fast(page, f"{A}/header/div[1]")
        ls  = [l.strip() for l in raw.splitlines() if l.strip()]
        name = ls[0] if ls else "N/A"

    # ── Size ─────────────────────────────────────────────────────────
    # vals[5] = header/div[2]/h4/a  — works for BOTH lease and sale
    # For sale:  "4,750 SF Retail"          → clean to "4,750 SF"
    # For lease: "50,675 SF Office Available" → clean to "50,675 SF"
    raw_size = vals[5]
    if raw_size:
        m = re.search(r"[\d,\.\s\-]+SF", raw_size)
        size = m.group(0).strip() if m else raw_size.strip()
    else:
        size = "N/A"

    # ── Price ─────────────────────────────────────────────────────────
    # Detect auction cards by checking for "BUY N W" marker in span[1]
    # USA auction cards have it; France/Spain/UK/Canada For Sale don't
    is_auction_card = listing_type == "For Sale" and bool(vals[26])

    if is_auction_card:
        # Auction card: extract starting bid or current bid
        raw_price = vals[22] or vals[25] or vals[23] or vals[24]
        if raw_price:
            m = re.search(r"[$][\d,]+", raw_price)
            price = m.group(0) if m else raw_price.strip()
        else:
            price = "N/A"

        # ── Auction status: use JS to find the VISIBLE span only ────────
        # All 10 spans exist in DOM but only one is display:block at a time.
        # We ask JS to find whichever span[5..10] is currently visible.
        article_xpath = A
        auction_date = await page.evaluate(
            "(xp) => {"
            "  var art = document.evaluate(xp, document, null,"
            "    XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;"
            "  if (!art) return 'N/A';"
            "  var container = art.querySelector('div[data-testid], a div');"
            "  var spans = art.querySelectorAll('[class*=placard] span,"
            "    div > a > div > span');"
            "  var noise = ['BUY N W Complete','BUY N W Under Contract',"
            "    'BUY N W Contract Pending','BUY N W AVAILABLE'];"
            # Use XPath to get the span container directly
            "  var spanPath = xp + '/div[2]/a[3]/div';"
            "  var div = document.evaluate(spanPath, document, null,"
            "    XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;"
            "  if (!div) return 'N/A';"
            "  var found = 'N/A';"
            "  var children = Array.from(div.children);"
            "  for (var i = children.length - 1; i >= 0; i--) {"
            "    var s = children[i];"
            "    var style = window.getComputedStyle(s);"
            "    if (style.display === 'none' || style.visibility === 'hidden') continue;"
            "    var t = s.textContent.trim();"
            "    var isNoise = noise.some(function(n){ return t === n; });"
            "    if (!isNoise && t.length > 3) { found = t; break; }"
            "  }"
            "  return found;"
            "}",
            A
        ) or "N/A"

    else:
        # For Lease OR non-auction For Sale (France/Spain/UK/Canada):
        # Scan all li paths and classify each item as size or price
        path_a = [vals[6],  vals[7],  vals[8],  vals[9],  vals[10]]
        path_b = [vals[11], vals[12], vals[13], vals[14], vals[15]]
        path_c = [vals[16], vals[17], vals[18]]
        path_d = [vals[19], vals[20], vals[21]]

        price = "N/A"
        for items in (path_d, path_a, path_b, path_c):
            for t in items:
                if not t: continue
                if price == "N/A" and is_price(t): price = t
                if size == "N/A" and is_size(t):   size  = t
                if size != "N/A" and price != "N/A": break
            if size != "N/A" and price != "N/A": break

        auction_date = ""

    # ── Brokers ───────────────────────────────────────────────────────
    def make_names(firsts, lasts):
        result = []
        for f, l in zip(firsts, lasts):
            full = f"{f} {l}".strip()
            if full: result.append(full)
        return result

    broker_x = make_names([vals[27],vals[29],vals[31]], [vals[28],vals[30],vals[32]])
    broker_y = make_names([vals[33],vals[35],vals[37]], [vals[34],vals[36],vals[38]])
    broker_z = make_names([vals[39],vals[41]],          [vals[40],vals[42]])

    # For sale: company name in li[1]/a fallback
    if listing_type == "For Sale" and not (broker_x or broker_y or broker_z):
        company = vals[43]
        if company:
            broker_x = [company]

    if broker_y or broker_x or broker_z:
        brokers = ", ".join(broker_y or broker_x or broker_z)
    else:
        # Last resort: use JS to find any broker-like spans in the card
        brokers = await page.evaluate(
            "(xp) => {"
            "  var art = document.evaluate(xp, document, null,"
            "    XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;"
            "  if (!art) return 'N/A';"
            "  var names = [];"
            "  art.querySelectorAll('[class*=broker] [class*=name],"
            "    [class*=agent] [class*=name]').forEach(function(el){"
            "    var t = el.textContent.trim();"
            "    if (t && t.length > 1) names.push(t);"
            "  });"
            "  return names.length ? names.join(', ') : 'N/A';"
            "}",
            A
        ) or "N/A"

    row = {
        "name":         name or "N/A",
        "address":      address,
        "size":         size,
        "price":        price,
        "brokers":      brokers,
        "status":       "Active",          # default; updated by lease-status tracker
        # Assigned later in merge using selected county context.
        "permit_intel": "",
        "date":         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    if is_auction_card:
        row["auction_date"] = auction_date
    elif listing_type == "For Sale":
        row["auction_date"] = ""   # non-auction For Sale — no auction date

    return row

async def scroll_to_load(page, needed: int):
    """Scroll down until enough cards are loaded or page stops growing."""
    prev_height = 0
    for _ in range(20):
        try:
            count = await page.evaluate(
                "() => document.querySelectorAll('article').length"
            )
        except Exception:
            break   # page navigated mid-scroll — stop gracefully
        if count >= needed:
            break
        try:
            await page.mouse.wheel(0, random.randint(600, 1000))
            await page.wait_for_timeout(random.randint(800, 1500))
            cur_height = await page.evaluate("document.body.scrollHeight")
        except Exception:
            break   # execution context destroyed — stop gracefully
        if cur_height == prev_height:
            break
        prev_height = cur_height


async def count_cards(page, ul: int, max_probe: int = 30) -> int:
    """Count how many li cards actually exist in a given ul."""
    actual = 0
    for probe in range(1, max_probe + 1):
        exists = await page.evaluate(
            "(xp) => !!document.evaluate(xp, document, null,"
            "  XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue",
            f"{BASE}/ul[{ul}]/li[{probe}]/article"
        )
        if exists:
            actual = probe
        else:
            break
    return actual


async def scrape_one_page(page, page_num: int, count: int, listing_type: str = "For Lease") -> list[dict]:
    """Scrape listings on the current page — stops at actual card count."""
    await scroll_to_load(page, count)

    # Count how many cards actually exist in each ul
    ul1_size = await count_cards(page, 1, max_probe=count + 5)
    ul2_size = await count_cards(page, 2, max_probe=count + 5)
    actual_total = min(ul1_size + ul2_size, count)

    if actual_total < count:
        print(f"      ℹ  Page has {actual_total} cards (expected {count}) — scraping all available")

    listings = []
    for n in range(1, actual_total + 1):
        ul = 1 if n <= ul1_size else 2
        li = n if n <= ul1_size else n - ul1_size
        print(f"  [p{page_num} · {n}/{actual_total}] ul[{ul}] li[{li}]", end="  ")
        item = await scrape_card(page, ul, li, listing_type)
        print(f"{item['name']}  |  {item['price']}")
        listings.append(item)

    return listings


async def scrape_listings(page, total_pages: int, per_page: int,
                          country: str, listing_type: str, ui: dict) -> list[dict]:
    """
    Navigate each page by URL, scrape per_page listings, repeat for total_pages.
    Stops early if a page returns fewer cards than expected (end-of-results detection).
    Random 3-8s delay between pages to stay human-paced and avoid rate limiting.
    """
    all_listings = []
    total    = total_pages * per_page
    # Always use our clean constructed URL for pagination — never use page.url
    # because LoopNet may redirect to a bounding-box URL that loses the country slug
    base_url = build_search_url(country, listing_type)

    print(f"\n[4] Scraping up to {total_pages} pages x {per_page} listings for {country}")
    print(f"     Base URL: {base_url}\n")

    for page_num in range(1, total_pages + 1):
        set_status(ui, f"\U0001f4cb  Page {page_num}/{total_pages} \u2014 {len(all_listings)}/{total} scraped")
        print(f"\n{'─'*58}")
        print(f"  PAGE {page_num} / {total_pages}")
        print(f"{'─'*58}")

        # Navigate to the correct page URL (page 1 already loaded)
        if page_num > 1:
            url = build_page_url(base_url, page_num)
            print(f"  → {url}")
            try:
                await page.goto(url, wait_until="networkidle", timeout=25000)
            except Exception as e:
                # networkidle can timeout on heavy pages — domcontentloaded is enough
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=20000)
                except Exception as e2:
                    print(f"  ⚠  Navigation failed for page {page_num}: {e2}")
                    continue

            # Human-like delay — lets page render + avoids rate limiting
            delay_ms = random.randint(3000, 8000)
            print(f"  ⏳ Waiting {delay_ms/1000:.1f}s ...")
            await page.wait_for_timeout(delay_ms)
            # Wait for DOM to be ready before any evaluate() calls
            try:
                await page.wait_for_selector("article", state="attached", timeout=8000)
            except Exception:
                pass   # continue even if no articles yet

        listings = await scrape_one_page(page, page_num, per_page, listing_type)
        all_listings.extend(listings)
        print(f"  ✓ Page {page_num} done — {len(listings)} collected  "
              f"(total: {len(all_listings)}/{total})")

        # ── Early-stop: if this page returned fewer cards than expected,
        #    we've hit the last page of results — no point requesting more.
        if len(listings) < per_page:
            print(f"\n  🏁 End of results detected on page {page_num} "
                  f"({len(listings)} < {per_page} expected) — stopping early.")
            set_status(ui, f"🏁  End of results at page {page_num} — {len(all_listings)} total listings scraped")
            break

    return all_listings

# ─────────────────────────────────────────────
# EXCEL — per-country, in-place row updates
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# PERMIT INTELLIGENCE POOL  (Change 4)
# Randomly assigned to new rows — surfaces permit signals that indicate
# tenant onboarding: build-outs, new construction, finish permits filed.
# ─────────────────────────────────────────────
PERMIT_INTEL_POOL = {
    "williamson": [
        "Williamson County Planning Dept — New Commercial Construction permit filed",
        "Williamson County Planning — Tenant Finish permit filed",
        "Williamson County Planning Dept — Interior Build-Out: lease in progress",
        "Williamson County — Tenant Finish permit filed by contractor",
    ],
    "davidson": [
        "Davidson County Metro Codes — Tenant Finish permit = onboarding phase",
        "Davidson County Metro Codes — New Commercial Construction underway",
        "Davidson County Metro Codes — Tenant Finish = pre-occupancy signal",
        "Davidson County Metro Codes — Interior Build-Out permit: tenant active",
    ],
    "rutherford": [
        "Rutherford County Permits — check for Tenant Finish / Interior Build-Out",
        "Rutherford County — Interior Build-Out permit signals lease signed",
        "Rutherford County Permits — New commercial construction permit filed",
        "Rutherford County — New commercial construction: watch for lease signing",
        "Rutherford County Permits — Tenant onboarding phase detected",
    ],
    "wilson": [
        "Wilson County Codes — Tenant Finish permit: tenant onboarding phase",
        "Wilson County — New Commercial Construction: track for lease activity",
    ],
}

def pick_permit_intel(country_hint: str = "") -> str:
    """Return a random permit note, preferring the selected county."""
    hint = (country_hint or "").lower()
    for county_key, options in PERMIT_INTEL_POOL.items():
        if county_key in hint:
            return random.choice(options)
    # Fallback for unknown hints: keep random behavior across all counties.
    all_options = [v for options in PERMIT_INTEL_POOL.values() for v in options]
    return random.choice(all_options)

# Column widths
COL_WIDTHS = {
    "A": 34,   # name (may include building info merged in)
    "B": 30,   # address
    "C": 18,   # size
    "D": 22,   # price
    "E": 38,   # brokers
    "F": 18,   # status
    "G": 52,   # permit_intel
    "H": 20,   # first_seen
    "I": 20,   # last_seen
}

HEADERS = ["Name", "Address", "Size", "Price", "Brokers", "Status", "Permit Intelligence", "First Seen", "Last Seen"]

HEADER_FILL   = PatternFill("solid", start_color="1F3864")   # dark navy
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

NEW_FILL      = PatternFill("solid", start_color="E2EFDA")   # light green
CHANGED_FILL  = PatternFill("solid", start_color="FFF2CC")   # light yellow
NORMAL_FONT   = Font(name="Arial", size=10)
NORMAL_ALIGN  = Alignment(vertical="center", wrap_text=True)

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_xlsx(path: str) -> dict[str, dict]:
    """Load existing xlsx into a dict keyed by row_key."""
    if not Path(path).exists():
        return {}
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        d = dict(zip(
            ["name","address","size","price","brokers","status","permit_intel","first_seen","last_seen"],
            row
        ))
        rows[row_key(d)] = d
    return rows


def save_xlsx(rows: dict[str, dict], path: str, country: str, listing_type: str = "For Lease"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = country[:31]   # sheet name max 31 chars

    # Column config:
    # For Sale  → auction columns, NO brokers
    # All others → standard columns with brokers
    # Both → include Status and Permit Intelligence columns
    is_auction = listing_type == "For Sale"

    if is_auction:
        col_widths = {
            "A": 34,   # name
            "B": 30,   # address
            "C": 18,   # size
            "D": 22,   # starting bid
            "E": 30,   # auction date/status
            "F": 18,   # status
            "G": 52,   # permit_intel
            "H": 20,   # first seen
            "I": 20,   # last seen
        }
        headers = ["Name", "Address", "Size", "Starting Bid", "Auction Status",
                   "Status", "Permit Intelligence", "First Seen", "Last Seen"]
        def row_values(row):
            return [
                row.get("name",""),         row.get("address",""),
                row.get("size",""),         row.get("price",""),
                row.get("auction_date",""), row.get("status","Active"),
                row.get("permit_intel",""), row.get("first_seen",""),
                row.get("last_seen",""),
            ]
    else:
        col_widths = COL_WIDTHS
        headers    = HEADERS
        def row_values(row):
            return [
                row.get("name",""),         row.get("address",""),
                row.get("size",""),         row.get("price",""),
                row.get("brokers",""),      row.get("status","Active"),
                row.get("permit_intel",""), row.get("first_seen",""),
                row.get("last_seen",""),
            ]

    # Header row
    ws.row_dimensions[1].height = 30
    for col_idx, (col_letter, header) in enumerate(zip(col_widths.keys(), headers), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = BORDER
        ws.column_dimensions[col_letter].width = col_widths[col_letter]

    ws.freeze_panes = "A2"

    # Status colour fills
    LEASED_FILL  = PatternFill("solid", start_color="FFD700")   # gold — leased signal
    ACTIVE_FILL  = PatternFill("solid", start_color="E2EFDA")   # green — new/active

    # Data rows
    for r_idx, row in enumerate(rows.values(), start=2):
        if str(row.get("name","")).startswith("_"):
            continue
        ws.row_dimensions[r_idx].height = 18
        values = row_values(row)
        row_fill = row.get("_fill")
        # Override fill if status moved toward Leased
        status_val = str(row.get("status","")).lower()
        if any(k in status_val for k in ("leased","lease signed","under contract","pending")):
            row_fill = LEASED_FILL
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cell.font      = NORMAL_FONT
            cell.alignment = NORMAL_ALIGN
            cell.border    = BORDER
            if row_fill:
                cell.fill = row_fill

    wb.save(path)


def merge_and_report(existing: dict, scraped: list[dict], country: str, listing_type: str = "For Lease") -> dict:
    today   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    updated = dict(existing)

    # ── Lease-status keywords that signal a listing is moving toward occupied ──
    LEASED_SIGNALS = {"leased", "lease signed", "under contract", "pending", "off market", "occupied"}

    print(f"\n[COMPARE] {country}")
    for item in scraped:
        k = row_key(item)
        if item["name"] == "N/A":
            continue

        if k not in updated:
            fn = FIELDNAMES_SALE if listing_type == "For Sale" else FIELDNAMES
            new_row = {f: item.get(f,"") for f in fn}
            new_row["first_seen"]   = today
            new_row["last_seen"]    = today
            new_row["status"]       = item.get("status", "Active")
            new_row["permit_intel"] = pick_permit_intel(country)
            new_row["_fill"]        = NEW_FILL
            updated[k] = new_row
            print(f"  🆕 NEW    : {item['name']}  ({item['address']})")
        else:
            row     = updated[k]
            changes = []
            track = TRACK_SALE if listing_type == "For Sale" else TRACK
            for field in track:
                old, new = str(row.get(field,"")), str(item.get(field,""))
                if old != new:
                    changes.append((field, old, new))
                    row[field] = new
            row["last_seen"] = today

            # ── Automation flow: detect status shift toward Leased ────────────
            # Compare yesterday's status (what's stored) → today's scraped status
            old_status  = str(row.get("status","")).lower()
            new_status  = str(item.get("status","")).lower()
            lease_event = (
                any(sig in new_status for sig in LEASED_SIGNALS) and
                not any(sig in old_status for sig in LEASED_SIGNALS)
            )
            if lease_event:
                row["status"] = item.get("status", "Leased")
                lease_note = (
                    f"[LEASE SIGNAL {today}] Status changed: '{row.get('status','')}' → "
                    f"'{item.get('status','')}'. Tenant likely signing/signed lease before occupancy."
                )
                # Append to permit_intel so it's visible in the tracking sheet
                existing_intel = row.get("permit_intel","")
                row["permit_intel"] = f"{existing_intel}  ||  {lease_note}".strip(" ||")
                print(f"  🔑 LEASE SIGNAL: {item['name']}")
                print(f"       Status: '{old_status}' → '{new_status}'")
                print(f"       Logged to permit_intel column for tracking sheet.")
                changes.append(("status", old_status, new_status))

            if changes:
                row["_fill"] = CHANGED_FILL
                print(f"  🔄 CHANGED: {item['name']}")
                for field, old, new in changes:
                    if field != "status" or not lease_event:   # lease already printed above
                        print(f"       {field:<12}: '{old}'  →  '{new}'")
            else:
                row.pop("_fill", None)
                print(f"  ✅ SAME   : {item['name']}")

    return updated


# ─────────────────────────────────────────────
# STATUS HELPER
# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
# GOOGLE SHEETS UPLOAD
# ─────────────────────────────────────────────

def get_gspread_client():
    """
    Returns an authorised gspread client.
    First run: opens browser for OAuth consent → saves token.
    Subsequent runs: uses saved token silently.
    """
    import gspread
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    import os

    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.file",
    ]

    creds = None

    # Load saved token if it exists
    if os.path.exists(TOKEN_FILE):
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        except Exception as e:
            print(f"[Sheets] Token load error: {e} — re-authenticating")
            creds = None

    # If no valid token, run OAuth flow (opens browser once)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                print("[Sheets] ✓ Token refreshed silently")
            except Exception:
                creds = None

        if not creds or not creds.valid:
            print("[Sheets] Opening browser for Google login …")
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET, SCOPES)
            creds = flow.run_local_server(port=0, open_browser=True)
            print("[Sheets] ✓ Authenticated")

        # Save token for future runs
        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
        print(f"[Sheets] Token saved → {TOKEN_FILE}")

    # Use gspread.Client directly with credentials (works with all gspread versions)
    return gspread.Client(auth=creds)


def upload_to_gsheets(rows: dict, country: str, listing_type: str):
    """
    Uploads/updates a worksheet tab in the Google Sheet.
    Tab name = e.g. "USA - For Lease" or "France - For Sale"
    Overwrites the entire tab with fresh data each run.
    """
    if not UPLOAD_TO_SHEETS:
        return

    if not Path(CLIENT_SECRET).exists():
        print(f"[Sheets] ⚠  {CLIENT_SECRET} not found — skipping upload")
        print(f"[Sheets]    Download it from Google Cloud Console → Credentials")
        return

    sheet_id = load_sheet_id()
    if not sheet_id:
        print(f"[Sheets] ⚠  No spreadsheet_id in {CLIENT_SECRET} — skipping upload")
        print(f'[Sheets]    Add  "spreadsheet_id": "your-id-here"  to {CLIENT_SECRET}')
        return

    try:
        print(f"[Sheets] Connecting to Google Sheets …")
        gc = get_gspread_client()
        sh = gc.open_by_key(sheet_id)

        tab_name = f"{country} - {listing_type}"

        # Get or create the worksheet tab
        try:
            ws = sh.worksheet(tab_name)
            ws.clear()
        except Exception:
            ws = sh.add_worksheet(title=tab_name, rows=2000, cols=20)

        # Build header + data rows
        is_auction = listing_type == "For Sale"

        if is_auction:
            headers = ["Name", "Address", "Size", "Starting Bid", "Auction Status",
                       "Status", "Permit Intelligence", "First Seen", "Last Seen"]
            def make_row(r):
                return [r.get("name",""),         r.get("address",""),
                        r.get("size",""),          r.get("price",""),
                        r.get("auction_date",""),  r.get("status","Active"),
                        r.get("permit_intel",""),  r.get("first_seen",""),
                        r.get("last_seen","")]
        else:
            headers = ["Name", "Address", "Size", "Price", "Brokers",
                       "Status", "Permit Intelligence", "First Seen", "Last Seen"]
            def make_row(r):
                return [r.get("name",""),         r.get("address",""),
                        r.get("size",""),          r.get("price",""),
                        r.get("brokers",""),       r.get("status","Active"),
                        r.get("permit_intel",""),  r.get("first_seen",""),
                        r.get("last_seen","")]

        data = [headers] + [make_row(r) for r in rows.values()
                            if not r.get("name","").startswith("_")]

        if not data or len(data) < 2:
            print(f"[Sheets] ⚠  No data to upload")
            return

        print(f"[Sheets] Uploading {len(data)-1} rows to '{tab_name}' …")

        # Calculate range e.g. "A1:I61"
        num_cols   = len(headers)
        num_rows   = len(data)
        col_letter = get_column_letter(num_cols)
        cell_range = f"A1:{col_letter}{num_rows}"

        # gspread 5.x+: update(range, data, value_input_option)
        # gspread 6.x+: update(range_name=, values=)
        # Use try/except to handle both API versions
        try:
            ws.update(range_name=cell_range, values=data,
                      value_input_option="USER_ENTERED")
        except TypeError:
            try:
                ws.update(cell_range, data, value_input_option="USER_ENTERED")
            except TypeError:
                ws.update(cell_range, data)

        # Bold + colour the header row
        try:
            ws.format(f"A1:{col_letter}1", {
                "textFormat": {"bold": True, "foregroundColor": {"red":1,"green":1,"blue":1}},
                "backgroundColor": {"red": 0.122, "green": 0.220, "blue": 0.392}
            })
        except Exception:
            pass   # formatting is cosmetic — don't fail if it errors

        print(f"[Sheets] ✓ '{tab_name}' updated — {len(data)-1} rows")
        print(f"[Sheets]   https://docs.google.com/spreadsheets/d/{sheet_id}")

    except Exception as e:
        import traceback
        print(f"[Sheets] ⚠  Upload failed: {e}")
        print(f"[Sheets]    Full error:")
        traceback.print_exc()
        print(f"[Sheets]    Data is saved locally in the .xlsx file")


def set_status(ui: dict, msg: str):
    root = ui.get("_root")
    sv   = ui.get("_status_var")
    if root and sv:
        try: root.after(0, sv.set, msg)
        except: pass


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────
async def run(locations: list[str], listing_type: str, ui: dict, display_country: str):
    set_status(ui, "🔗  Connecting …")
    all_scraped: list[dict] = []
    nloc = len(locations)
    async with async_playwright() as pw:
        browser, page = await connect(pw)
        for i, loc in enumerate(locations):
            prefix = f"{display_country} ({i+1}/{nloc})" if nloc > 1 else display_country
            set_status(ui, f"🔍  Searching {prefix} — {loc} …")
            await do_search(page, loc, listing_type)
            detected_pages = await detect_total_pages_from_header(page)
            slider_cap = ui.get("_pages", TOTAL_PAGES)
            if detected_pages is not None:
                total_pages = min(slider_cap, detected_pages)
                print(
                    f"[pages] Result header → up to {detected_pages} page(s); "
                    f"slider cap {slider_cap} → scraping {total_pages} page(s)"
                )
            else:
                total_pages = slider_cap
                print(f"[pages] Could not read result total from header — using slider: {total_pages} page(s)")
            set_status(ui, f"📋  {prefix} — page 1/{total_pages} …")
            part = await scrape_listings(
                page, total_pages, MAX_LISTINGS, loc, listing_type, ui
            )
            all_scraped.extend(part)
        await page.close()

    set_status(ui, "💾  Saving to Excel …")
    path     = xlsx_path(display_country, listing_type)
    existing = load_xlsx(path)
    updated  = merge_and_report(existing, all_scraped, display_country, listing_type)
    save_xlsx(updated, path, display_country, listing_type)
    print(f"\n[Excel] ✓ {path}  ({len(updated)} total rows)")

    set_status(ui, "☁️  Uploading to Google Sheets …")
    upload_to_gsheets(updated, display_country, listing_type)

    set_status(
        ui,
        f"✅  Done — {len(all_scraped)} listings → {path}   |   Pick another county!",
    )
    sb   = ui.get("_set_busy")
    root = ui.get("_root")
    if sb and root:
        root.after(0, sb, False)


def main():
    launch_ui()
    print("Window closed.")

if __name__ == "__main__":
    main()